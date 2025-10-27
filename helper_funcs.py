#helper_funcs.py
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.styles import Font, Alignment, Border, Protection, PatternFill
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.formatting import Rule
import pandas as pd
from copy import copy 
from datetime import datetime
import re
import os

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

def populate_underpayment_rows(lead_sheet, num_rows_to_add):
    """
    Populates the underpayment rows in the TP3 lead sheet.

    Args:
        lead_sheet: The lead sheet.
        num_rows_to_add: The number of rows added to TP3.2.
    """

    start_row = 5
    end_row = start_row + num_rows_to_add
    index_end_row = 14 + num_rows_to_add  # last row of index

    for col_letter in range(column_letter_to_index("A"), column_letter_to_index("AS") + 1):
        excel_col_letter = column_index_to_letter(col_letter)

        for row in range(start_row, end_row):
            # Construct the dynamic formula
             formula = (
                f'=IFERROR(INDEX(\'TP.3.2_Lockdown Period\'!{excel_col_letter}$15:{excel_col_letter}$1015; '
                f'MATCH(ROW()-4; \'TP.3.2_Lockdown Period\'!$AT$15:$AT$1015; 0)); "")'
            )
             
             lead_sheet[f'{excel_col_letter}{row}'].value = formula

def add_table_copy_formula(
    target_sheet: Worksheet,
    start_cell: str,
    table_copy_text: str,
    payments_sheet_name: str,
    num_rows_to_add: int
):
    """
    Inserts a formula to copy a table dynamically from `payments_sheet_2` (A:AS) to the `target_sheet`.

    Args:
        target_sheet (Worksheet): The worksheet where the formula will be inserted.
        start_cell (str): The starting cell where the table should be copied (e.g., "A13").
        table_copy_text (str): The text to search for where the table copy formula should be placed.
        payments_sheet_name (str): The name of the sheet containing the original table.
        num_rows_to_add (int): The number of rows to copy.
    """
    # Step 1: Locate the cell with `table_copy_text`
    table_copy_cell = None
    for row in target_sheet.iter_rows():
        for cell in row:
            if cell.value == table_copy_text:
                table_copy_cell = cell
                break
        if table_copy_cell:
            break

    if not table_copy_cell:
        raise ValueError(f"'{table_copy_text}' not found in the sheet.")

    # Step 2: Parse the start cell
    start_column = ''.join([char for char in start_cell if char.isalpha()])  # Extract column (e.g., "A" from "A13")
    start_row = int(''.join([char for char in start_cell if char.isdigit()]))  # Extract row number (e.g., 13 from "A13")

    # Step 3: Calculate the last row using `num_rows_to_add`
    last_row = start_row + num_rows_to_add - 1  # Adjust range to include the starting row

    # Step 4: Construct the formula to copy the range from `payments_sheet_2`
    source_range = f"'{payments_sheet_name}'!A{start_row}:AS{last_row}"
    formula = f"=ARRAYFORMULA({source_range})"

    # Step 5: Insert the formula into the cell below `table_copy_text`
    target_cell = target_sheet.cell(row=table_copy_cell.row + 1, column=table_copy_cell.column)
    target_cell.value = formula



def add_conclusion_formula(
    target_sheet: Worksheet, 
    start_cell: str, 
    conclusion_text: str, 
    true_cell: str, 
    true_cond_cell: str, 
    false_cell: str, 
    num_rows_to_add: int
):
    """
    Add a formula below the 'Conclusion' cell in the sheet.

    The formula:
    - Returns `true_cell` if all values in the range are "a".
    - Returns `true_cond_cell` if at least one value in the range is "a".
    - Returns `false_cell` if no values in the range are "a".

    Args:
        target_sheet (Worksheet): The target worksheet.
        start_cell (str): The starting cell for checking the range (e.g., "U13").
        conclusion_text (str): The text to search for in the sheet where the conclusion is defined.
        true_cell (str): The cell reference to use if all values are "a" (e.g., "Data!B3").
        true_cond_cell (str): The cell reference to use if at least one value is "a" (e.g., "Data!B4").
        false_cell (str): The cell reference to use if no values are "a" (e.g., "Data!B5").
        num_rows_to_add (int): The number of rows added starting from `start_cell`.
    """
    # Step 1: Locate the cell with "Conclusion"
    conclusion_cell = None
    for row in target_sheet.iter_rows():
        for cell in row:
            if cell.value == conclusion_text:
                conclusion_cell = cell
                break
        if conclusion_cell:
            break

    if not conclusion_cell:
        raise ValueError(f"'{conclusion_text}' not found in the sheet.")

    # Step 2: Parse the start cell
    start_column = ''.join([char for char in start_cell if char.isalpha()])  # Extract column (e.g., "U" from "U13")
    start_row = int(''.join([char for char in start_cell if char.isdigit()]))  # Extract row number (e.g., 13 from "U13")

    # Step 3: Calculate the last row using `num_rows_to_add`
    last_row = start_row + num_rows_to_add - 1  # Adjust range to include the starting row

    # Step 4: Create the formula
    range_to_check = f"{start_column}{start_row}:{start_column}{last_row}"
    formula = (
        f'=IF(COUNTIF({range_to_check}, "a")=ROWS({range_to_check}), {true_cell}, '
        f'IF(COUNTIF({range_to_check}, "a")>0, {true_cond_cell}, {false_cell}))'
    )

    # Step 5: Insert the formula into the cell below "Conclusion"
    target_cell = target_sheet.cell(row=conclusion_cell.row + 1, column=conclusion_cell.column)
    target_cell.value = formula



def load_data_file(data_file_path):
    """
    Load the data file and return the workbook and sheet.

    Args:
        data_file_path (str): Path to the data file.

    Returns:
        tuple: A tuple containing the loaded workbook and the first sheet from the workbook.
    """
    data_wb = load_workbook(data_file_path, data_only=True)
    data_sheet = data_wb.worksheets[0]
    return data_wb, data_sheet

def load_working_paper(working_paper_path, sh_n):
    """
    Load the working paper template, unlock all sheets, and return the workbook and lead sheet.

    Args:
        working_paper_path (str): Path to the working paper template.
        sh_n (int): Index of the lead sheet (0-based).
        
    Returns:
        tuple: A tuple containing the loaded working paper workbook and the specified lead sheet.
        
    Raises:
        Exception: If the password for unlocking the sheets is incorrect or the workbook cannot be loaded.
    """
    password = "wp2"
    
    # Suppress the data validation warning by temporarily redirecting stderr
    import warnings
    import sys
    from io import StringIO
    
    # Temporarily suppress warnings during workbook loading
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        # Load the workbook as regular .xlsx without VBA macros
        working_paper_wb = load_workbook(working_paper_path, data_only=False)
    
    # Iterate over all worksheets in the workbook and unlock them
    for sheet in working_paper_wb.worksheets:
        if sheet.protection.sheet:
            sheet.protection.set_password(password)  # Unlock sheet with password, if provided
            sheet.protection.sheet = False  # Disable protection
    
    # Get the specified lead sheet by index
    lead_sheet = working_paper_wb.worksheets[sh_n]
    
    return working_paper_wb, lead_sheet

def convert_to_dataframe(data_sheet):
    """
    Convert the sheet data to a DataFrame, applying filtering conditions.

    Args:
        data_sheet (Worksheet): The sheet from which data will be extracted.

    Returns:
        pd.DataFrame: A DataFrame containing the filtered data based on predefined conditions.
    """
    # Convert the sheet data to a DataFrame
    df = pd.DataFrame(
        data_sheet.iter_rows(values_only=True, min_row=2),
        columns=[cell.value for cell in data_sheet[1]]
    )
    
    # Apply the filtering conditions
    df = df[(df['PAYMENT_STATUS_ID'] == 3) & (df['PAYMENTMEDIUMID'] == 2) & (df['BANK_PAY_AMOUNT'] != 0 )]
    
    return df

def get_column_indexes(data_sheet):
    """
    Extract the column indexes based on the headings in the data sheet.

    Args:
        data_sheet (Worksheet): The sheet from which column indexes are to be extracted.

    Returns:
        dict: A dictionary where keys are column headings and values are their respective indexes.
    """
    headings = {cell.value: idx for idx, cell in enumerate(data_sheet[1], start=1)}
    return headings

def extract_tradename_uif(data_sheet, headings):
    """
    Extract the 'TRADENAME' and 'UIFREFERENCENUMBER' from the data sheet.

    Args:
        data_sheet (Worksheet): The sheet from which the 'TRADENAME' and 'UIFREFERENCENUMBER' will be extracted.
        headings (dict): A dictionary containing the column indexes of the relevant headings.

    Returns:
        tuple: A tuple containing the extracted 'TRADENAME' and 'UIFREFERENCENUMBER'.
    """
    tradename = next(data_sheet.iter_rows(min_row=2, min_col=headings["TRADENAME"], max_col=headings["TRADENAME"], values_only=True))[0]
    uif_reference = next(data_sheet.iter_rows(min_row=2, min_col=headings["UIFREFERENCENUMBER"], max_col=headings["UIFREFERENCENUMBER"], values_only=True))[0]
    return tradename, uif_reference

def extract_shutdown_periods(data_sheet, headings):
    """
    Extract and return unique shutdown periods as a formatted string in chronological order.

    Args:
        data_sheet (Worksheet): The sheet from which shutdown periods will be extracted.
        headings (dict): A dictionary containing the column indexes of the relevant headings.

    Returns:
        str: A string representing the unique shutdown periods in chronological order.
    """
    shutdown_from_col = headings["SHUTDOWN_FROM"]
    shutdown_till_col = headings["SHUTDOWN_TILL"]
    periods = set()  # Use a set to store unique periods (to avoid duplicates)

    # Define possible date formats for parsing
    date_formats = [
        "%Y-%m-%d %H:%M:%S",  # Standard datetime format
        "%Y-%m-%d",           # ISO date format
        "%d/%m/%Y",           # European date format
        "%m/%d/%Y",           # US date format
        "%d-%b-%Y",           # Day-Month-Year with abbreviated month name
        "%d %B %Y",           # Day-Month-Year with full month name
    ]

    # Loop through each row in the data sheet and extract shutdown periods
    for row in data_sheet.iter_rows(min_row=2, values_only=True):
        shutdown_from = row[shutdown_from_col - 1]
        shutdown_till = row[shutdown_till_col - 1]

        if shutdown_from and shutdown_till:
            from_date, till_date = None, None

            # Try parsing the dates using the available formats
            for fmt in date_formats:
                try:
                    if not from_date:
                        from_date = datetime.strptime(str(shutdown_from), fmt)
                    if not till_date:
                        till_date = datetime.strptime(str(shutdown_till), fmt)
                except ValueError:
                    continue

            # If parsing was successful, add the period to the set
            if from_date and till_date:
                periods.add((from_date, till_date))

    # Sort the periods by the 'from_date'
    sorted_periods = sorted(periods, key=lambda x: x[0])

    # Convert the sorted periods into a string format
    periods_str = ", ".join([f"{period[0].strftime('%d %B %Y')} to {period[1].strftime('%d %B %Y')}" for period in sorted_periods])
    return periods_str

def create_output_directory(output_directory, tradename, wp_n, uif_reference=None, data_file_path=None, template_paths=None, create_folders_only=False):
    """
    Create a folder structure in the output directory for saving processed files and return the full processed file path.

    Args:
        output_directory (str): Path to the directory where the output folders will be created.
        tradename (str): The tradename to be used in the output folder name.
        wp_n (int): An integer identifying the specific type of testing or working paper.
        uif_reference (str): The UIF reference number to be used in the parent folder name.
        data_file_path (str): Path to the original data file to copy to UIF DATAFILE folder.
        template_paths (list): List of template file paths to copy to AUDIT REPORTING TEMPLATES folder.
        create_folders_only (bool): If True, only create the folder structure and copy files, don't return a working paper path.

    Returns:
        str: The full file path where the processed file will be saved, or the parent folder path if create_folders_only=True.
    
    Raises:
        ValueError: If `wp_n` is not in the range [1, 2, 3, 4], a ValueError will be raised.
    """
    # Ensure the tradename and UIF reference are safe for use in file paths
    safe_tradename = "".join(c if c.isalnum() or c in (" ", "_", "-") else "_" for c in tradename).strip()
    safe_uif_ref = "".join(c if c.isalnum() or c in (" ", "_", "-") else "_" for c in (uif_reference or "")).strip()
    
    # Create parent folder name: {UIF Reg Number} - {Company name/tradename}
    if safe_uif_ref:
        parent_folder_name = f"{safe_uif_ref} - {safe_tradename}"
    else:
        parent_folder_name = f"UIF_REF - {safe_tradename}"
    
    # Create the main parent folder
    parent_folder = os.path.join(output_directory, parent_folder_name)
    if not os.path.exists(parent_folder):
        os.makedirs(parent_folder)
    
    # Create the 4 subfolders
    subfolders = [
        "AUDIT REPORTING TEMPLATES",
        "AUDIT WORKING PAPERS", 
        "INFORMATION FROM EMPLOYER",
        "UIF DATAFILE"
    ]
    
    for subfolder in subfolders:
        subfolder_path = os.path.join(parent_folder, subfolder)
        if not os.path.exists(subfolder_path):
            os.makedirs(subfolder_path)
    
    # Copy data file to UIF DATAFILE folder if provided
    if data_file_path and os.path.exists(data_file_path):
        import shutil
        data_filename = os.path.basename(data_file_path)
        uif_datafile_path = os.path.join(parent_folder, "UIF DATAFILE", data_filename)
        try:
            shutil.copy2(data_file_path, uif_datafile_path)
        except Exception as e:
            print(f"Warning: Could not copy data file to UIF DATAFILE folder: {e}")
    
    # Copy report templates to AUDIT REPORTING TEMPLATES folder with UIF reference naming
    if safe_uif_ref:
        import shutil
        script_dir = os.path.dirname(os.path.abspath(__file__))
        report_templates_dir = os.path.join(script_dir, "TEMPLATES", "Report_Templates")
        audit_templates_path = os.path.join(parent_folder, "AUDIT REPORTING TEMPLATES")
        
        if os.path.exists(report_templates_dir):
            for filename in os.listdir(report_templates_dir):
                if filename.endswith(('.docx', '.xlsx', '.pdf')):  # Common document formats
                    source_path = os.path.join(report_templates_dir, filename)
                    # Create new filename with UIF reference: {File Name} - {UIF REG NUMBER}
                    file_name_without_ext = os.path.splitext(filename)[0]
                    file_extension = os.path.splitext(filename)[1]
                    new_filename = f"{file_name_without_ext} - {safe_uif_ref}{file_extension}"
                    dest_path = os.path.join(audit_templates_path, new_filename)
                    try:
                        shutil.copy2(source_path, dest_path)
                    except Exception as e:
                        print(f"Warning: Could not copy report template {filename} to AUDIT REPORTING TEMPLATES folder: {e}")
    
    # If only creating folders, return the parent folder path
    if create_folders_only:
        return parent_folder
    
    # Determine folder name and file name based on wp_n
    if wp_n == 1:
        folder_name = "TP.1_Compliance and Existence Testing"
        file_name = f"TP.1_Compliance and Existence Testing_{safe_uif_ref}.xlsx"
    elif wp_n == 2:
        folder_name = "TP.2_Employment Verification Testing"
        file_name = f"TP.2_Employment Verification Testing_{safe_uif_ref}.xlsx"
    elif wp_n == 3:
        folder_name = "TP.3_Payment Verification"
        file_name = f"TP.3_Payment Verification_{safe_uif_ref}.xlsx"
    elif wp_n == 4:
        folder_name = "TP.4_Confirmation of UIF Contributions"
        file_name = f"TP.4_Confirmation of UIF Contributions_{safe_uif_ref}.xlsx"
    else:
        raise ValueError("Invalid wp_n value. Must be 1, 2, 3, or 4.")

    # Create the processed file path in the AUDIT WORKING PAPERS subfolder
    audit_working_papers_folder = os.path.join(parent_folder, "AUDIT WORKING PAPERS")
    tp_x_folder = os.path.join(audit_working_papers_folder, folder_name)
    if not os.path.exists(tp_x_folder):
        os.makedirs(tp_x_folder)
    processed_file_path = os.path.join(tp_x_folder, file_name)
    
    return processed_file_path

def copy_cell_style_and_formula(source_cell, target_cell, target_row):
    """
    Copy both the styles, formulas, and data validation from a source cell to a target cell, adjusting for row references.

    Args:
        source_cell (Cell): The source cell containing the style and formula.
        target_cell (Cell): The target cell where the style and formula will be copied to.
        target_row (int): The target row for the formula row reference adjustment.

    Returns:
        None

    Raises:
        Exception: If there is an error in copying the style or formula.
    """
    try:
        # Copy the style
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font) if source_cell.font else None
            target_cell.alignment = copy(source_cell.alignment) if source_cell.alignment else None
            target_cell.border = copy(source_cell.border) if source_cell.border else None
            target_cell.fill = copy(source_cell.fill) if source_cell.fill else None
            target_cell.number_format = source_cell.number_format
            target_cell.protection = copy(source_cell.protection) if source_cell.protection else None

        # Note: Data validation is handled automatically by Excel when rows are inserted
        # We don't need to manually copy data validation as it's preserved by Excel's built-in functionality

        # Copy the formula if it exists, adjust for the new row reference
        if source_cell.value and isinstance(source_cell.value, str) and source_cell.value.startswith('='):
            formula = source_cell.value
            # Update the row references in the formula to match the target row
            updated_formula = re.sub(r'(\$?[A-Za-z]+)(\d+)', lambda m: f"{m.group(1)}{target_row}" if m.group(2) != str(target_row) else m.group(0), formula)
            target_cell.value = updated_formula
        elif source_cell.value:
            target_cell.value = source_cell.value  # Copy value if it's not a formula

    except Exception as e:
        print(f"Error copying cell style and formula: {e}")
        
def insert_rows(employee_sheet, num_rows_to_add, insert_start_row):
    """
    Insert rows into the sheet at the specified location.

    Args:
        employee_sheet (Worksheet): The worksheet where the rows will be inserted.
        num_rows_to_add (int): The number of rows to insert.
        insert_start_row (int): The row number where the new rows will be inserted.

    Returns:
        None

    Raises:
        Exception: If there is an error while inserting rows.
    """
    try:
        employee_sheet.insert_rows(insert_start_row, amount=num_rows_to_add)
    except Exception as e:
        print(f"Error while inserting rows: {e}")

def copy_formatting(employee_sheet, START__ROW, num_rows_to_add, source_cell_n):
    """
    Copy the formatting (styles, formulas, row height) from a source row to a set of target rows.

    Args:
        employee_sheet (Worksheet): The worksheet where the formatting will be applied.
        START__ROW (int): The starting row to begin copying formatting from.
        num_rows_to_add (int): The number of rows that will receive the copied formatting.
        source_cell_n (int): The source row that contains the original formatting to copy.

    Returns:
        None

    Raises:
        Exception: If there is an error while copying the formatting.
    """
    # Note: Data validation ranges are automatically adjusted by Excel when rows are inserted
    # We don't need to manually adjust them as Excel handles this automatically
    
    for row_idx in range(START__ROW, START__ROW + num_rows_to_add):
        for col_idx in range(1, employee_sheet.max_column + 1):
            try:
                col_letter = get_column_letter(col_idx)
                source_cell = employee_sheet[f"{col_letter}{source_cell_n}"]
                target_cell = employee_sheet[f"{col_letter}{row_idx}"]
                copy_cell_style_and_formula(source_cell, target_cell, row_idx)
                employee_sheet.row_dimensions[row_idx].height = employee_sheet.row_dimensions[source_cell_n].height
            except Exception as e:
                print(f"Error while copying cell formatting and formulas: {e}")

def reset_row_heights(sheet, reference_row, target_rows, hide_reference_row=False):
    """
    Reset the heights of target rows based on a reference row's height and optionally hide the reference row.

    Args:
        sheet (Worksheet): The sheet object where the rows will be adjusted.
        reference_row (int): The row number whose height will be used as a reference.
        target_rows (list): A list or range of row numbers to adjust.
        hide_reference_row (bool): Flag to optionally hide the reference row after adjusting.

    Returns:
        None
    """
    reference_height = sheet.row_dimensions[reference_row].height

    for row in target_rows:
        sheet.row_dimensions[row].height = reference_height

    if hide_reference_row:
        sheet.row_dimensions[reference_row].hidden = True

def unmerge_cells_in_range(sheet, start_row, end_row):
    """
    Unmerge cells within a specified row range and store their original boundaries, styles, and row heights.

    Args:
        sheet (Worksheet): The sheet object where the unmerging will take place.
        start_row (int): The starting row of the range to check for merged cells.
        end_row (int): The ending row of the range to check for merged cells.

    Returns:
        list: A list of tuples representing the original merged cell boundaries, styles, and row heights.
    """
    merged_cells_to_restore = []
    for merged_range in list(sheet.merged_cells.ranges):
        try:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            # Check if the merged cell overlaps with the target range
            if not (end_row < min_row or start_row > max_row):
                source_cell = sheet.cell(row=min_row, column=min_col)
                row_height = sheet.row_dimensions[min_row].height or sheet.row_dimensions.defaultRowHeight  # Capture row height
                merged_cells_to_restore.append(
                    (min_col, min_row, max_col, max_row, source_cell.alignment, row_height)
                )
                sheet.unmerge_cells(str(merged_range))
        except Exception as e:
            import traceback
            print(f"Error in unmerge_cells_in_range: {e}")
            traceback.print_exc()
    return merged_cells_to_restore

def reapply_merged_cells(sheet, merged_cells_to_restore, num_rows_to_add):
    """
    Reapply merged cells and restore their alignment and row heights after rows have been added.

    Args:
        sheet (Worksheet): The sheet object to modify.
        merged_cells_to_restore (list): List of tuples with the original merged cell boundaries, alignment, and row height.
        num_rows_to_add (int): The number of rows to adjust the merged cells' ranges by.

    Returns:
        None
    """
    for min_col, min_row, max_col, max_row, alignment, row_height in merged_cells_to_restore:
        try:
            adjusted_min_row = min_row + num_rows_to_add
            adjusted_max_row = max_row + num_rows_to_add
            sheet.merge_cells(
                start_row=adjusted_min_row, start_column=min_col,
                end_row=adjusted_max_row, end_column=max_col
            )

            # Apply the saved alignment to all cells in the merged range
            for row in range(adjusted_min_row, adjusted_max_row + 1):
                for col in range(min_col, max_col + 1):
                    cell = sheet.cell(row=row, column=col)
                    if alignment:
                       cell.alignment = alignment.copy()
                       
            # Reapply the row height to the adjusted minimum row
            sheet.row_dimensions[adjusted_min_row].height = row_height
        except Exception as e:
            print(f"Error while reapplying merged cells: {e}")

def validate_columns(data, required_columns):
    """
    Validate that all required columns are present in the dataset.

    Args:
        data (pd.DataFrame): A pandas DataFrame containing the data to validate.
        required_columns (list): A list of column names that must be present in the DataFrame.

    Raises:
        KeyError: If any required column is missing from the DataFrame.
    """
    missing_columns = [col for col in required_columns if col not in data.columns]
    if missing_columns:
        raise KeyError(f"Missing required columns: {', '.join(missing_columns)}")


def apply_conditional_formatting_general(employee_sheet, start_row, num_rows_to_add, columns_to_format, legend):
    """
    Apply conditional formatting to specific columns and rows to highlight empty cells and 'r' values.

    Args:
        employee_sheet (openpyxl.worksheet.worksheet.Worksheet): The sheet to which conditional formatting will be applied.
        start_row (int): The starting row number for formatting.
        num_rows_to_add (int): The number of rows to add formatting.
        columns_to_format (list): The list of columns (letters) to apply formatting to.
        legend (str): The legend column (e.g., 'L') to apply specific formatting for 'r' values.
    """
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    red_fill_legend = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for row_idx in range(start_row, start_row + num_rows_to_add):
        for col_letter in columns_to_format:
            cell_range = f"{col_letter}{row_idx}"
            rule_empty = CellIsRule(operator="equal", formula=['""'], stopIfTrue=True, fill=red_fill)
            employee_sheet.conditional_formatting.add(cell_range, rule_empty)

        cell_range_legend = f"{legend}{row_idx}"
        rule_r = CellIsRule(operator="equal", formula=['"r"'], stopIfTrue=True, fill=red_fill_legend)
        employee_sheet.conditional_formatting.add(cell_range_legend, rule_r)


def column_letter_to_index(column_letter):
    """
    Convert Excel-style column letters (e.g., 'A', 'AA') to a 1-based index.

    Args:
        column_letter (str): The column letter(s) to convert (e.g., 'A', 'Z', 'AA').

    Returns:
        int: The corresponding column index (1-based).
    """
    column_index = 0
    for char in column_letter:
        column_index = column_index * 26 + (ord(char.upper()) - ord('A') + 1)
    return column_index


def column_index_to_letter(column_index):
    """
    Convert a 1-based column index to Excel-style column letters.

    Args:
        column_index (int): The 1-based column index to convert (e.g., 1, 26, 28).

    Returns:
        str: The corresponding column letter(s) (e.g., 'A', 'Z', 'AB').
    """
    column_letter = ""
    while column_index > 0:
        column_index -= 1
        column_letter = chr(column_index % 26 + ord('A')) + column_letter
        column_index //= 26
    return column_letter


def save_working_paper(working_paper_wb, processed_file_path):
    """
    Save the modified working paper to the specified location as .xlsx.

    Args:
        working_paper_wb (openpyxl.workbook.workbook.Workbook): The workbook object to save.
        processed_file_path (str): The file path where the workbook should be saved.
    """
    working_paper_wb.save(processed_file_path)


def get_unique_id_count(datasheet, column_name="IDNUMBER"):
    """
    Counts the number of unique ID numbers in the specified column of the datasheet.

    Args:
        datasheet (pd.DataFrame): The input datasheet as a Pandas DataFrame.
        column_name (str): The name of the column to analyze for unique ID numbers.

    Returns:
        int: The count of unique ID numbers in the column.
    
    Raises:
        ValueError: If the column does not exist in the datasheet.
    """
    datasheet = convert_to_dataframe(datasheet)
    if column_name not in datasheet.columns:
        raise ValueError(f"Column '{column_name}' not found in the datasheet.")
    
    unique_ids = datasheet[column_name].nunique()
    return unique_ids


def get_bank_pay_amount_sum(datasheet):
    """
    Returns the sum of the 'BANK_PAY_AMOUNT' column in the specified datasheet, rounded to 2 decimal points.

    Args:
        datasheet (pd.DataFrame): The input datasheet as a Pandas DataFrame.

    Returns:
        float: The sum of the 'BANK_PAY_AMOUNT' column, rounded to 2 decimal points.
    
    Raises:
        ValueError: If the 'BANK_PAY_AMOUNT' column is missing from the datasheet.
    """
    datasheet = convert_to_dataframe(datasheet)
    if 'BANK_PAY_AMOUNT' not in datasheet.columns:
        raise ValueError("Column 'BANK_PAY_AMOUNT' not found in the datasheet.")
    
    total_amount = datasheet['BANK_PAY_AMOUNT'].sum()
    return round(total_amount, 2)


def get_company_info(data_file_path):
    """
    Extracts company information from the given data file.

    Args:
        data_file_path (str): The file path to the data file.

    Returns:
        tuple: A tuple containing the following information:
            - company_name (str): The name of the company.
            - uif_ref (str): The UIF reference for the company.
            - periods_claimed (list): The periods claimed by the company.
            - number_of_employees (int): The number of employees.
            - total_amount_claimed (float): The total amount claimed by the company.
    """
    data_wb, data_sheet = load_data_file(data_file_path)
    convert_to_dataframe(data_sheet)
    headings = get_column_indexes(data_sheet)
    company_name, uif_ref = extract_tradename_uif(data_sheet, headings)
    periods_claimed = extract_shutdown_periods(data_sheet, headings)
    number_of_employees = get_unique_id_count(data_sheet)
    total_amount_claimed = get_bank_pay_amount_sum(data_sheet)
    
    return company_name, uif_ref, periods_claimed, number_of_employees, total_amount_claimed

def update_formulas_after_row_insertion(sheet, insert_start_row, num_rows_added):
    """
    Update all existing formulas in the sheet to account for inserted rows.
    
    When rows are inserted, existing formulas that reference cells below the insertion point
    need to have their row references adjusted upward by the number of rows inserted.
    
    Args:
        sheet (Worksheet): The worksheet containing the formulas to update.
        insert_start_row (int): The row number where rows were inserted.
        num_rows_added (int): The number of rows that were inserted.
    """
    try:
        # Iterate through all cells in the sheet
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value
                    updated_formula = adjust_formula_references(formula, insert_start_row, num_rows_added)
                    if updated_formula != formula:
                        cell.value = updated_formula
    except Exception as e:
        print(f"Error updating formulas after row insertion: {e}")

def adjust_formula_references(formula, insert_start_row, num_rows_added):
    """
    Adjust cell references in a formula to account for inserted rows.
    
    Args:
        formula (str): The original formula string.
        insert_start_row (int): The row number where rows were inserted.
        num_rows_added (int): The number of rows that were inserted.
    
    Returns:
        str: The updated formula with adjusted cell references.
    """
    try:
        # More comprehensive pattern to match cell references
        # This handles: A1, $B$5, C10:D15, A1:B5, $A$1:$B$5, etc.
        cell_ref_pattern = r'(\$?[A-Za-z]+\$?\d+(?::\$?[A-Za-z]+\$?\d+)?)'
        
        def adjust_cell_reference(match):
            cell_ref = match.group(1)
            
            # Handle range references (e.g., A1:B5)
            if ':' in cell_ref:
                start_cell, end_cell = cell_ref.split(':')
                adjusted_start = adjust_single_cell_reference(start_cell, insert_start_row, num_rows_added)
                adjusted_end = adjust_single_cell_reference(end_cell, insert_start_row, num_rows_added)
                return f"{adjusted_start}:{adjusted_end}"
            else:
                return adjust_single_cell_reference(cell_ref, insert_start_row, num_rows_added)
        
        updated_formula = re.sub(cell_ref_pattern, adjust_cell_reference, formula)
        return updated_formula
    
    except Exception as e:
        print(f"Error adjusting formula references: {e}")
        return formula

def adjust_single_cell_reference(cell_ref, insert_start_row, num_rows_added):
    """
    Adjust a single cell reference to account for inserted rows.
    
    Args:
        cell_ref (str): The cell reference (e.g., A1, $B$5).
        insert_start_row (int): The row number where rows were inserted.
        num_rows_added (int): The number of rows that were inserted.
    
    Returns:
        str: The adjusted cell reference.
    """
    try:
        # Extract column and row parts more robustly
        # Handle both relative and absolute references: A1, $A$1, $A1, A$1
        col_part = ''
        row_part = ''
        dollar_before_col = False
        dollar_before_row = False
        
        i = 0
        while i < len(cell_ref):
            char = cell_ref[i]
            if char == '$':
                if not col_part:  # Dollar before column
                    dollar_before_col = True
                else:  # Dollar before row
                    dollar_before_row = True
            elif char.isalpha():
                col_part += char
            elif char.isdigit():
                row_part += char
            i += 1
        
        if not row_part:
            return cell_ref  # No row reference to adjust
        
        row_num = int(row_part)
        
        # Only adjust if the referenced row is at or below the insertion point
        if row_num >= insert_start_row:
            new_row_num = row_num + num_rows_added
            # Reconstruct the cell reference with proper dollar signs
            result = ''
            if dollar_before_col:
                result += '$'
            result += col_part
            if dollar_before_row:
                result += '$'
            result += str(new_row_num)
            return result
        else:
            return cell_ref
    
    except Exception as e:
        print(f"Error adjusting single cell reference: {e}")
        return cell_ref


def create_folder_structure_for_all_working_papers(output_directory, tradename, uif_reference, data_file_path, template_paths):
    """
    Create the folder structure for all working papers and copy necessary files.
    
    Args:
        output_directory (str): Path to the directory where the output folders will be created.
        tradename (str): The tradename to be used in the output folder name.
        uif_reference (str): The UIF reference number to be used in the parent folder name.
        data_file_path (str): Path to the original data file to copy to UIF DATAFILE folder.
        template_paths (list): List of template file paths to copy to AUDIT REPORTING TEMPLATES folder.
    
    Returns:
        str: The path to the AUDIT WORKING PAPERS subfolder where all working papers will be saved.
    """
    # Create the folder structure and copy files
    parent_folder = create_output_directory(
        output_directory, tradename, 1, uif_reference, data_file_path, template_paths, create_folders_only=True
    )
    
    # Return the path to the AUDIT WORKING PAPERS subfolder
    audit_working_papers_folder = os.path.join(parent_folder, "AUDIT WORKING PAPERS")
    return audit_working_papers_folder


def get_working_paper_path_for_all_processing(audit_working_papers_folder, tradename, wp_n, uif_reference):
    """
    Get the path for a working paper when processing all working papers together.
    
    Args:
        audit_working_papers_folder (str): Path to the AUDIT WORKING PAPERS subfolder.
        tradename (str): The tradename to be used in the filename.
        wp_n (int): An integer identifying the specific type of testing or working paper.
        uif_reference (str): The UIF reference to be used in the filename.
    
    Returns:
        str: The full file path where the working paper will be saved.
    
    Raises:
        ValueError: If `wp_n` is not in the range [1, 2, 3, 4], a ValueError will be raised.
    """
    # Ensure the tradename and UIF reference are safe for use in file paths
    safe_tradename = "".join(c if c.isalnum() or c in (" ", "_", "-") else "_" for c in tradename).strip()
    safe_uif_ref = "".join(c if c.isalnum() or c in (" ", "_", "-") else "_" for c in (uif_reference or "")).strip()
    
    # Determine folder name and file name based on wp_n
    if wp_n == 1:
        folder_name = "TP.1_Compliance and Existence Testing"
        file_name = f"TP.1_Compliance and Existence Testing_{safe_uif_ref}.xlsx"
    elif wp_n == 2:
        folder_name = "TP.2_Employment Verification Testing"
        file_name = f"TP.2_Employment Verification Testing_{safe_uif_ref}.xlsx"
    elif wp_n == 3:
        folder_name = "TP.3_Payment Verification"
        file_name = f"TP.3_Payment Verification_{safe_uif_ref}.xlsx"
    elif wp_n == 4:
        folder_name = "TP.4_Confirmation of UIF Contributions"
        file_name = f"TP.4_Confirmation of UIF Contributions_{safe_uif_ref}.xlsx"
    else:
        raise ValueError("Invalid wp_n value. Must be 1, 2, 3, or 4.")

    # Create folder path within the AUDIT WORKING PAPERS folder
    tp_x_folder = os.path.join(audit_working_papers_folder, folder_name)
    if not os.path.exists(tp_x_folder):
        os.makedirs(tp_x_folder)

    # Create the processed file path
    processed_file_path = os.path.join(tp_x_folder, file_name)
    
    return processed_file_path
