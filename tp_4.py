#tp_4.py
from helper_funcs import (
    load_data_file, 
    load_working_paper, 
    get_column_indexes, 
    extract_tradename_uif, 
    create_output_directory, 
    save_working_paper,
    get_working_paper_path_for_all_processing,
    unmerge_cells_in_range,
    reapply_merged_cells
)
from openpyxl import load_workbook
from datetime import datetime

def process_files(data_file_path, working_paper_path, consultant_name, output_directory):
    """
    Main function to process the data file and update the working paper with the extracted information.
    
    This function:
    - Loads the data and working paper files.
    - Extracts necessary details such as tradename, UIF reference, shutdown periods, and current date.
    - Populates the lead sheet of the working paper with the extracted data.
    - Creates an output directory for saving the processed file.
    - Saves the updated working paper to the specified output directory.

    Args:
        data_file_path (str): Path to the source data file (e.g., Excel file).
        working_paper_path (str): Path to the working paper file to be updated.
        consultant_name (str): Name of the consultant responsible for the working paper.
        output_directory (str): Path to the directory where the processed working paper will be saved.

    Returns:
        str: Path to the saved working paper after processing and updates.
    
    Raises:
        Exception: If there are any errors loading the files or extracting data.
    """
    try:
        # Load the data file (Excel workbook and the first sheet)
        data_wb, data_sheet = load_data_file(data_file_path)

        # Load the working paper template (Excel workbook and lead sheet)
        working_paper_wb, lead_sheet = load_working_paper(working_paper_path, sh_n=0)

        # Extract column indexes from the data sheet
        headings = get_column_indexes(data_sheet)
        
        # Extract tradename and UIF reference from the data sheet
        tradename, uif_reference = extract_tradename_uif(data_sheet, headings)

        # Use fixed string for periods
        periods_str = "Lockdown Periods"

        # Get the current date for record-keeping
        current_date = datetime.now().strftime("%Y-%m-%d")
        
        # Populate the lead sheet with the extracted data
        populate_working_paper(lead_sheet, tradename, uif_reference, periods_str, current_date, consultant_name)

        # Create an output directory and get the processed file path
        processed_file_path = create_output_directory(output_directory, tradename, wp_n=4, uif_reference=uif_reference, data_file_path=data_file_path, template_paths=[working_paper_path])

        # Save the modified working paper
        save_working_paper(working_paper_wb, processed_file_path)

        return processed_file_path

    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        raise


def process_files_for_all_processing(data_file_path, working_paper_path, consultant_name, audit_working_papers_folder):
    """
    Main function to process the data file and update the working paper with the extracted information.
    This version is used when processing all working papers together.
    
    This function:
    - Loads the data and working paper files.
    - Extracts necessary details such as tradename, UIF reference, shutdown periods, and current date.
    - Populates the lead sheet of the working paper with the extracted data.
    - Saves the updated working paper to the pre-created AUDIT WORKING PAPERS folder.

    Args:
        data_file_path (str): Path to the source data file (e.g., Excel file).
        working_paper_path (str): Path to the working paper file to be updated.
        consultant_name (str): Name of the consultant responsible for the working paper.
        audit_working_papers_folder (str): Path to the AUDIT WORKING PAPERS subfolder.

    Returns:
        str: Path to the saved working paper after processing and updates.
    
    Raises:
        Exception: If there are any errors loading the files or extracting data.
    """
    try:
        # Load the data file (Excel workbook and the first sheet)
        data_wb, data_sheet = load_data_file(data_file_path)

        # Load the working paper template (Excel workbook and lead sheet)
        working_paper_wb, lead_sheet = load_working_paper(working_paper_path, sh_n=0)

        # Extract column indexes from the data sheet
        headings = get_column_indexes(data_sheet)
        
        # Extract tradename and UIF reference from the data sheet
        tradename, uif_reference = extract_tradename_uif(data_sheet, headings)

        # Use fixed string for periods
        periods_str = "Lockdown Periods"

        # Get the current date for record-keeping
        current_date = datetime.now().strftime("%Y-%m-%d")
        
        # Populate the lead sheet with the extracted data
        populate_working_paper(lead_sheet, tradename, uif_reference, periods_str, current_date, consultant_name)

        # Get the processed file path in the pre-created folder structure
        processed_file_path = get_working_paper_path_for_all_processing(audit_working_papers_folder, tradename, wp_n=4, uif_reference=uif_reference)

        # Save the modified working paper
        save_working_paper(working_paper_wb, processed_file_path)

        return processed_file_path

    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        raise


def populate_working_paper(lead_sheet, tradename, uif_reference, periods_str, current_date, consultant_name):
    """
    Populates the working paper with extracted data.

    This function updates specific cells in the provided lead sheet of the working paper
    with the extracted data, including tradename, UIF reference, shutdown periods, current
    date, and the consultant's name. It safely handles merged cells by unmerging them
    before writing and then remerging them after.

    Args:
        lead_sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet object to populate with data.
        tradename (str): The tradename to be inserted into the working paper.
        uif_reference (str): The UIF reference to be inserted into the working paper.
        periods_str (str): The shutdown periods to be inserted into the working paper.
        current_date (str): The current date to be inserted into the working paper.
        consultant_name (str): The name of the consultant to be inserted into the working paper.

    Returns:
        None
    """
    try:
        # First, unmerge any cells in the company info area (rows 1-4, columns B and F)
        # This covers the range where company info is typically inserted
        merged_cells_to_restore = unmerge_cells_in_range(lead_sheet, start_row=1, end_row=4)
        
        # Insert the extracted data into specific cells in the working paper
        lead_sheet["B1"].value = tradename
        lead_sheet["B2"].value = uif_reference
        lead_sheet["B4"].value = periods_str
        lead_sheet["F3"].value = current_date
        lead_sheet["F1"].value = consultant_name  # Insert the consultant's name into F1
        
        # Reapply any merged cells that were temporarily unmerged
        # Note: We pass 0 as num_rows_added since we're not adding rows, just unmerging for writing
        reapply_merged_cells(lead_sheet, merged_cells_to_restore, 0)
        
        print(f"DEBUG: Successfully populated company info - Company: {tradename}, UIF: {uif_reference}")
        
    except Exception as e:
        print(f"ERROR in populate_working_paper: {e}")
        print(f"ERROR DETAILS - Function: populate_working_paper, Error: {type(e).__name__}")
        raise
