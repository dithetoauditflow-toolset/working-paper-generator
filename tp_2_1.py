#tp_2_1.py
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill

# Constants
START_ROW = 13  # Starting row for employee data insertion

def aggregate_data_2_1(data):
    """
    Aggregates employee data by ID number.

    This function groups the input data by "IDNUMBER" and calculates aggregate values for the fields such as 
    first name, last name, employment start date, termination date, bank pay amount, leave income, and monthly salary.
    It then sorts the data by last name.

    Args:
        data (pandas.DataFrame): The input data to aggregate, containing employee records.

    Returns:
        pandas.DataFrame: The aggregated data with the calculated values, sorted by last name.
    """
    # Aggregate the data
    aggregated_data = data.groupby("IDNUMBER").agg({
        "FIRSTNAME": "first",
        "LASTNAME": "first",
        "EMPLOYMENTSTARTDATE": "first",
        "TERMINATIONDATE": "min",
        "BANK_PAY_AMOUNT": "sum",
        "LEAVE_INCOME": "sum",
        "MONTHLY_SALARY": "first"
    }).reset_index()
    
    # Sort the rows by LASTNAME
    aggregated_data = aggregated_data.sort_values(by="LASTNAME")
    
    return aggregated_data

def populate_sheet_2_1(employee_sheet, aggregated):
    """
    Populates the employee sheet with aggregated data.

    This function takes the aggregated employee data and populates the corresponding cells in the employee sheet.
    It begins inserting data starting from the defined start row (START_ROW) and continues row by row.

    Args:
        employee_sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet object to populate with data.
        aggregated (pandas.DataFrame): The aggregated data to populate into the sheet.

    Returns:
        None
    """
    current_row = START_ROW
    for _, row in aggregated.iterrows():
        try:
            for col_idx, value in enumerate(row, start=1):
                adjusted_col_idx = col_idx if col_idx < 4 else col_idx + 1
                col_letter = get_column_letter(adjusted_col_idx)
                employee_sheet[f"{col_letter}{current_row}"].value = value
            current_row += 1
        except Exception as e:
            print(f"Error while populating aggregated data: {e}")
